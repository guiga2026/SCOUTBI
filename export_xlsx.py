"""Export all PostgreSQL tables and views to an Excel workbook."""

import argparse
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import inspect, text

from sports_bi.app.database import engine

MAX_EXCEL_ROWS = 1_048_576
MAX_SHEET_NAME = 31


def safe_sheet_name(name: str, used: set[str]) -> str:
    base = "".join(char if char not in "[]:*?/\\" else "_" for char in name)[:MAX_SHEET_NAME] or "sheet"
    candidate = base
    suffix = 1
    while candidate in used:
        ending = f"_{suffix}"
        candidate = f"{base[:MAX_SHEET_NAME - len(ending)]}{ending}"
        suffix += 1
    used.add(candidate)
    return candidate


def export_workbook(output: Path) -> int:
    inspector = inspect(engine)
    schema = "public" if engine.dialect.name == "postgresql" else None
    entities = [("table", name) for name in inspector.get_table_names(schema=schema)]
    entities += [("view", name) for name in inspector.get_view_names(schema=schema)]
    entities.sort(key=lambda item: (item[0], item[1]))

    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    readme.append(["Sports BI export"])
    readme.append(["Generated at (UTC)", datetime.now(timezone.utc).isoformat()])
    readme.append(["Source", "PostgreSQL public schema"])
    readme.append([])
    readme.append(["Type", "Entity", "Rows", "Sheet"])
    for cell in readme[5]:
        cell.font = Font(bold=True)

    used_sheets = {"README"}
    total_rows = 0
    with engine.connect() as connection:
        for entity_type, entity_name in entities:
            sheet = workbook.create_sheet(safe_sheet_name(entity_name, used_sheets))
            escaped_name = entity_name.replace(chr(34), chr(34) * 2)
            qualified_name = f'public."{escaped_name}"' if schema else f'"{escaped_name}"'
            result = connection.execute(text(f"SELECT * FROM {qualified_name}"))
            columns = list(result.keys())
            sheet.append(columns)
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            rows = 0
            for row in result:
                sheet.append(list(row))
                rows += 1
                total_rows += 1
                if rows >= MAX_EXCEL_ROWS - 1:
                    break
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for index, column in enumerate(columns, start=1):
                sheet.column_dimensions[get_column_letter(index)].width = min(max(len(str(column)) + 2, 12), 35)
            readme.append([entity_type, entity_name, rows, sheet.title])

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return total_rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Exporta tabelas e views do Sports BI para XLSX")
    parser.add_argument("--output", type=Path, default=Path("/tmp/sports_bi_export.xlsx"))
    args = parser.parse_args()
    total_rows = export_workbook(args.output)
    print(f"XLSX gerado: {args.output} ({total_rows} linhas)", flush=True)


if __name__ == "__main__":
    main()
